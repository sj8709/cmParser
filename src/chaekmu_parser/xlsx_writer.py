"""
ParsedDocument → XLSX 템플릿 채움.

템플릿(`templates/chaekmu_template.xlsx`) 복제 후:
  1. `임원_템플릿` 시트를 임원 수만큼 복제
  2. 복제본 시트명 = 직책명 (\\n→', ', 31자 절단, 중복 시 + 성명)
  3. 기본 데이터 영역 초기화 후 §1/§2/§3 채움
  4. 가변 행(회의체/책무/관리의무)은 뒤에서 앞 순서로 삽입·삭제 후 서식 상속

레이아웃은 템플릿 `설정` 시트 스펙을 그대로 따름.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from chaekmu_parser.models import (
    Committee,
    Executive,
    Obligation,
    ParsedDocument,
    Responsibility,
)

TEMPLATE_NAME_SETTING = "설정"
TEMPLATE_NAME_EXEC = "임원_템플릿"
SHEET_NAME_MAX = 31
TEXT_FMT = "@"


@dataclass(frozen=True)
class TemplateLayout:
    """임원_템플릿 시트의 초기 행 좌표."""
    # §1 임원정보 (수직 레이아웃, 고정)
    pos_row: int = 5
    name_row: int = 6
    title_row: int = 7
    appt_row: int = 8
    concur_yn_row: int = 9
    concur_dtl_row: int = 10
    dept_row: int = 11

    # §1 회의체
    committee_header_row: int = 14
    committee_data_start: int = 15
    committee_data_base: int = 4
    committee_footer_row: int = 19  # 회의체 데이터 기본 끝(18) + 1

    # §2 책무 개요
    resp_summary_row: int = 22
    resp_assign_row: int = 23
    resp_header_row: int = 26
    resp_data_start: int = 27
    resp_data_base: int = 5
    resp_footer_row: int = 32

    # §3 관리의무
    oblig_header_row: int = 36
    oblig_data_start: int = 37
    oblig_data_base: int = 9
    oblig_footer_row: int = 46


LAYOUT = TemplateLayout()


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------
def write(parsed: ParsedDocument, template_path: Path, output_path: Path) -> None:
    wb = load_workbook(str(template_path))
    if TEMPLATE_NAME_EXEC not in wb.sheetnames:
        raise ValueError(f"템플릿에 '{TEMPLATE_NAME_EXEC}' 시트가 없음")

    planned_names = _plan_sheet_names(parsed.executives)

    for exec_, sheet_name in zip(parsed.executives, planned_names):
        ws = wb.copy_worksheet(wb[TEMPLATE_NAME_EXEC])
        ws.title = sheet_name
        _fill_executive_sheet(ws, exec_)

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# 시트명 계획 (중복 대응 포함)
# ---------------------------------------------------------------------------
def _plan_sheet_names(executives: list[Executive]) -> list[str]:
    base_names = [_sheet_name_from_position(e.position) for e in executives]
    counts: dict[str, int] = {}
    for n in base_names:
        counts[n] = counts.get(n, 0) + 1
    seen: dict[str, int] = {}
    result: list[str] = []
    for exec_, base in zip(executives, base_names):
        if counts[base] > 1:
            combined = f"{base} - {exec_.name}"
            result.append(combined[:SHEET_NAME_MAX])
        else:
            seen[base] = seen.get(base, 0) + 1
            if seen[base] == 1:
                result.append(base)
            else:
                result.append(f"{base}_{seen[base]}"[:SHEET_NAME_MAX])
    return result


def _sheet_name_from_position(position: str) -> str:
    name = position.replace("\n", ", ").strip()
    # Excel 시트명 금지 문자
    for bad in ("\\", "/", "*", "[", "]", ":", "?"):
        name = name.replace(bad, "")
    return name[:SHEET_NAME_MAX]


# ---------------------------------------------------------------------------
# 임원 시트 작성
# ---------------------------------------------------------------------------
def _fill_executive_sheet(ws: Worksheet, exec_: Executive) -> None:
    actual_committee = len(exec_.committees) if exec_.committees else 1  # 0개 → 1행 N/A
    actual_resp = max(len(exec_.responsibilities), 1)
    actual_oblig = max(len(exec_.obligations), 1)

    delta_committee = actual_committee - LAYOUT.committee_data_base
    delta_resp = actual_resp - LAYOUT.resp_data_base
    delta_oblig = actual_oblig - LAYOUT.oblig_data_base

    # 1. 데이터 영역 초기화 — 각주/헤더 행은 보존
    _clear_data_area(ws)

    # 2. 행 조정 — 뒤에서 앞 순서. 각 호출은 원 footer 위치 기준 (상위는 아직 unchanged)
    _adjust_block(
        ws,
        footer_row=LAYOUT.oblig_footer_row,
        last_data_row=LAYOUT.oblig_data_start + LAYOUT.oblig_data_base - 1,
        delta=delta_oblig,
    )
    _adjust_block(
        ws,
        footer_row=LAYOUT.resp_footer_row,
        last_data_row=LAYOUT.resp_data_start + LAYOUT.resp_data_base - 1,
        delta=delta_resp,
    )
    _adjust_block(
        ws,
        footer_row=LAYOUT.committee_footer_row,
        last_data_row=LAYOUT.committee_data_start + LAYOUT.committee_data_base - 1,
        delta=delta_committee,
    )

    # 3. 조정 후 최종 좌표 (상위 섹션 삽입/삭제가 하위 행 번호를 모두 밀어올림)
    committee_start = LAYOUT.committee_data_start
    committee_end = committee_start + actual_committee - 1

    resp_summary_row = LAYOUT.resp_summary_row + delta_committee
    resp_assign_row = LAYOUT.resp_assign_row + delta_committee
    resp_start = LAYOUT.resp_data_start + delta_committee
    resp_end = resp_start + actual_resp - 1

    oblig_start = LAYOUT.oblig_data_start + delta_committee + delta_resp
    oblig_end = oblig_start + actual_oblig - 1

    # 4. 데이터 작성
    _write_executive_info(ws, exec_)
    _write_committees(ws, exec_.committees, committee_start, committee_end)
    _set(ws, resp_summary_row, 2, exec_.responsibility_summary)
    _set(ws, resp_assign_row, 2, exec_.assign_date)
    _write_responsibilities(ws, exec_.responsibilities, resp_start, resp_end)
    _write_obligations(ws, exec_.obligations, oblig_start, oblig_end)

    # 5. 모든 커스텀 행 높이 제거 → Excel이 열 때 내용 기반 auto-fit
    _reset_row_heights(ws)


def _reset_row_heights(ws: Worksheet) -> None:
    """모든 row_dimensions의 height를 None으로 설정해 Excel auto-fit 유도."""
    for rd in ws.row_dimensions.values():
        rd.height = None


# ---------------------------------------------------------------------------
# 초기화
# ---------------------------------------------------------------------------
# 템플릿에서 값/서식을 보존해야 할 행:
#   - 14/26/36: 리스트 헤더 (회의체/책무/관리의무 컬럼 헤더 텍스트)
#   - 19/32/46: 각주 텍스트 (B열 법령 근거 각주)
_PROTECTED_ROWS = frozenset({14, 19, 26, 32, 36, 46})


def _clear_data_area(ws: Worksheet) -> None:
    """B5:D46 데이터 영역을 value=None으로 초기화. 헤더/각주 행은 건너뜀 (서식/텍스트 보존)."""
    for row_idx in range(5, LAYOUT.oblig_footer_row + 1):
        if row_idx in _PROTECTED_ROWS:
            continue
        for col in range(2, 5):
            ws.cell(row=row_idx, column=col).value = None


# ---------------------------------------------------------------------------
# 행 조정
# ---------------------------------------------------------------------------
def _adjust_block(
    ws: Worksheet, footer_row: int, last_data_row: int, delta: int
) -> None:
    """
    한 섹션의 데이터 영역 크기를 delta만큼 조정.

    delta > 0: footer 바로 앞에 delta개 행 삽입 (마지막 데이터행의 서식 상속).
    delta < 0: 마지막 데이터행 근처에서 -delta개 삭제.
    delta == 0: 변화 없음.
    """
    if delta > 0:
        _insert_rows_with_style(
            ws, insert_at=footer_row, count=delta, template_row=last_data_row
        )
    elif delta < 0:
        delete_at = last_data_row + delta + 1
        ws.delete_rows(delete_at, -delta)


def _insert_rows_with_style(
    ws: Worksheet, insert_at: int, count: int, template_row: int
) -> None:
    """`insert_at` 앞에 `count` 개 빈 행 삽입 후 template_row의 서식을 복사."""
    ws.insert_rows(insert_at, count)
    # insert 후 template_row가 insert_at 이상이었다면 아래로 count만큼 밀림
    actual_tpl = template_row if template_row < insert_at else template_row + count
    max_col = ws.max_column
    for new_row in range(insert_at, insert_at + count):
        for col in range(1, max_col + 1):
            src = ws.cell(row=actual_tpl, column=col)
            dst = ws.cell(row=new_row, column=col)
            if src.has_style:
                dst._style = copy(src._style)


# ---------------------------------------------------------------------------
# 데이터 작성 헬퍼
# ---------------------------------------------------------------------------
def _set(ws: Worksheet, row: int, col: int, value: str | None) -> None:
    cell = ws.cell(row=row, column=col)
    cell.value = value if value else None
    cell.number_format = TEXT_FMT
    # 기존 alignment 속성(수평/수직 정렬 등) 유지하면서 wrap_text만 활성화
    existing = cell.alignment
    cell.alignment = Alignment(
        horizontal=existing.horizontal,
        vertical=existing.vertical or "top",
        wrap_text=True,
        shrink_to_fit=existing.shrink_to_fit,
        indent=existing.indent,
        text_rotation=existing.text_rotation,
    )


def _write_executive_info(ws: Worksheet, e: Executive) -> None:
    _set(ws, LAYOUT.pos_row, 2, e.position.replace("\n", ", "))
    _set(ws, LAYOUT.name_row, 2, e.name)
    _set(ws, LAYOUT.title_row, 2, e.title)
    _set(ws, LAYOUT.appt_row, 2, e.appointed_date)
    _set(ws, LAYOUT.concur_yn_row, 2, e.concurrent_yn)
    _set(ws, LAYOUT.concur_dtl_row, 2, e.concurrent_detail or "N/A")
    _set(ws, LAYOUT.dept_row, 2, e.departments)


def _write_committees(
    ws: Worksheet, committees: list[Committee], start: int, end: int
) -> None:
    if not committees:
        _set(ws, start, 1, "N/A")
        return
    for idx, c in enumerate(committees):
        r = start + idx
        if r > end:
            break
        _set(ws, r, 1, c.name)
        _set(ws, r, 2, c.role)
        _set(ws, r, 3, c.cycle)
        _set(ws, r, 4, c.matters)


def _write_responsibilities(
    ws: Worksheet, responsibilities: list[Responsibility], start: int, end: int
) -> None:
    if not responsibilities:
        return
    for idx, r in enumerate(responsibilities):
        row = start + idx
        if row > end:
            break
        _set(ws, row, 1, r.category)
        _set(ws, row, 2, r.details[0] if r.details else "")
        _set(ws, row, 3, r.raw_law_reg)


def _write_obligations(
    ws: Worksheet, obligations: list[Obligation], start: int, end: int
) -> None:
    if not obligations:
        return
    for idx, o in enumerate(obligations):
        row = start + idx
        if row > end:
            break
        _set(ws, row, 1, o.type)
        _set(ws, row, 2, o.category)
        _set(ws, row, 3, "\n".join(f"- {item}" for item in o.items))
