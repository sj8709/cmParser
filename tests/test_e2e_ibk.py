"""
IBK E2E 회귀 — DOCX 원본 ↔ 생성 XLSX 왕복 검증.

회귀 기준:
  A. 시트 구조: 설정 + 임원_템플릿 + 임원별 시트 N개
  B. 섹션 행 위치: delta 기반 좌표 계산과 실제 셀 내용 일치
  C. 각주/헤더: 원문 보존 (빈 셀 아님)
  D. 데이터 정합: ParsedDocument 필드값이 XLSX의 올바른 셀에 그대로 기록

expected.xlsx는 수동편집본이므로 비교 대상 아님.
"""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from chaekmu_parser.extractors.docx_extractor import DocxExtractor
from chaekmu_parser.normalizer import normalize
from chaekmu_parser.xlsx_writer import write

ROOT = Path(__file__).parent.parent
FIXTURE = ROOT / "fixtures/ibk/input.docx"
TEMPLATE = ROOT / "templates/chaekmu_template.xlsx"
pytestmark = pytest.mark.skipif(
    not (FIXTURE.exists() and TEMPLATE.exists()),
    reason="fixture/template not present",
)


# ---------------------------------------------------------------------------
# 공용 픽스처: 파이프라인 1회 실행 → (wb, parsed) 반환
# ---------------------------------------------------------------------------
@pytest.fixture(scope="module")
def pipeline(tmp_path_factory):
    parsed = normalize(DocxExtractor().extract(FIXTURE))
    out = tmp_path_factory.mktemp("e2e") / "out.xlsx"
    write(parsed, TEMPLATE, out)
    wb = load_workbook(str(out))
    return wb, parsed


def _sheet_name(position: str) -> str:
    n = position.replace("\n", ", ")
    for bad in ("\\", "/", "*", "[", "]", ":", "?"):
        n = n.replace(bad, "")
    return n.strip()[:31]


def _position_of(e) -> int:
    """임원별 delta 계산: (dc, dr, do)."""
    dc = (len(e.committees) or 1) - 4
    dr = max(len(e.responsibilities), 1) - 5
    do = max(len(e.obligations), 1) - 9
    return dc, dr, do


# ---------------------------------------------------------------------------
# A. 시트 구조
# ---------------------------------------------------------------------------
def test_sheet_structure(pipeline):
    wb, parsed = pipeline
    assert wb.sheetnames[0] == "설정"
    assert wb.sheetnames[1] == "임원_템플릿"
    exec_sheets = wb.sheetnames[2:]
    assert len(exec_sheets) == len(parsed.executives) == 9


def test_sheet_names_match_positions(pipeline):
    wb, parsed = pipeline
    exec_sheets = wb.sheetnames[2:]
    for e, name in zip(parsed.executives, exec_sheets):
        assert name == _sheet_name(e.position), (
            f"시트명 불일치: parsed={e.position!r} -> expected {_sheet_name(e.position)!r}, got {name!r}"
        )


# ---------------------------------------------------------------------------
# B. 섹션 행 위치 (delta 기반)
# ---------------------------------------------------------------------------
def test_section_anchors_per_executive(pipeline):
    """각 임원 시트에서 §1/§2/§3 제목과 ▼ 헤더가 delta 계산 위치에 존재."""
    wb, parsed = pipeline
    for e in parsed.executives:
        ws = wb[_sheet_name(e.position)]
        dc, dr, do = _position_of(e)

        # §1 제목 row 3 (불변)
        assert ws["A3"].value and "임원" in str(ws["A3"].value)
        # ▼ 주관 회의체 row 13 (불변 — committee 위)
        assert "회의체" in str(ws["A13"].value or "")
        # §2 제목 row 20 + dc
        assert "책무" in str(ws.cell(20 + dc, 1).value or ""), e.position
        # §3 제목 row 33 + dc + dr
        assert "관리의무" in str(ws.cell(33 + dc + dr, 1).value or ""), e.position


# ---------------------------------------------------------------------------
# C. 각주/헤더 보존
# ---------------------------------------------------------------------------
def test_list_headers_preserved(pipeline):
    """회의체/책무/관리의무 컬럼 헤더 텍스트가 각 시트에 그대로 남아있음."""
    wb, parsed = pipeline
    for e in parsed.executives:
        ws = wb[_sheet_name(e.position)]
        dc, dr, _ = _position_of(e)
        # 회의체 헤더 row 14
        assert "회의체명" in str(ws.cell(14, 1).value or ""), e.position
        assert "위원장" in str(ws.cell(14, 2).value or ""), e.position
        # 책무 내용 헤더 row 26 + dc
        assert str(ws.cell(26 + dc, 1).value or "").strip() == "책무"
        assert "세부" in str(ws.cell(26 + dc, 2).value or "")
        # 관리의무 헤더 row 36 + dc + dr
        assert "구분" in str(ws.cell(36 + dc + dr, 1).value or "")
        assert "책무명" in str(ws.cell(36 + dc + dr, 2).value or "")


def test_footnotes_preserved(pipeline):
    """§1/§2/§3 각주의 B열 법령 근거 원문이 delta 만큼 밀린 위치에 유지."""
    wb, parsed = pipeline
    for e in parsed.executives:
        ws = wb[_sheet_name(e.position)]
        dc, dr, do = _position_of(e)
        f1 = 19 + dc
        f2 = 32 + dc + dr
        f3 = 46 + dc + dr + do
        for r, sec in [(f1, "§1"), (f2, "§2"), (f3, "§3")]:
            txt = str(ws.cell(r, 2).value or "")
            assert txt.startswith("주") or "주 " in txt[:10], (
                f"{e.position} {sec} 각주 누락 row {r}: {txt[:50]!r}"
            )


# ---------------------------------------------------------------------------
# D. 데이터 정합: ParsedDocument → XLSX 셀 정확 기록
# ---------------------------------------------------------------------------
def test_executive_info_written_correctly(pipeline):
    wb, parsed = pipeline
    for e in parsed.executives:
        ws = wb[_sheet_name(e.position)]
        assert str(ws["B5"].value) == e.position.replace("\n", ", ")
        assert ws["B6"].value == e.name
        assert ws["B7"].value == e.title
        assert ws["B8"].value == e.appointed_date
        assert ws["B9"].value == e.concurrent_yn
        assert ws["B11"].value == e.departments


def test_committees_written_at_correct_rows(pipeline):
    wb, parsed = pipeline
    for e in parsed.executives:
        ws = wb[_sheet_name(e.position)]
        if not e.committees:
            assert ws.cell(15, 1).value == "N/A"
            continue
        for idx, c in enumerate(e.committees):
            row = 15 + idx
            assert ws.cell(row, 1).value == c.name
            assert ws.cell(row, 2).value == c.role
            assert ws.cell(row, 3).value == c.cycle
            assert ws.cell(row, 4).value == c.matters


def test_resp_summary_and_data_written_at_correct_rows(pipeline):
    wb, parsed = pipeline
    for e in parsed.executives:
        ws = wb[_sheet_name(e.position)]
        dc, _, _ = _position_of(e)
        assert ws.cell(22 + dc, 2).value == e.responsibility_summary
        assert ws.cell(23 + dc, 2).value == e.assign_date
        for idx, r in enumerate(e.responsibilities):
            row = 27 + dc + idx
            assert ws.cell(row, 1).value == r.category
            expected_detail = r.details[0] if r.details else ""
            assert (ws.cell(row, 2).value or "") == expected_detail
            assert (ws.cell(row, 3).value or "") == (r.raw_law_reg or "")


def test_obligations_written_with_correct_type_and_items(pipeline):
    wb, parsed = pipeline
    for e in parsed.executives:
        ws = wb[_sheet_name(e.position)]
        dc, dr, _ = _position_of(e)
        start = 37 + dc + dr
        for idx, o in enumerate(e.obligations):
            row = start + idx
            assert ws.cell(row, 1).value == o.type, e.position
            assert ws.cell(row, 2).value == o.category
            c_val = str(ws.cell(row, 3).value or "")
            # 각 세부 항목이 '- ' 접두사 + 개행 규칙으로 병합되었는지
            for item in o.items:
                assert item in c_val, f"{e.position} oblig[{idx}] item 누락"
            assert c_val.startswith("- "), e.position


# ---------------------------------------------------------------------------
# E. 출력 파일 재로드 가능성 (손상 방지)
# ---------------------------------------------------------------------------
def test_output_file_reloadable_without_errors(pipeline):
    wb, _ = pipeline
    assert len(wb.sheetnames) >= 3
    for name in wb.sheetnames:
        ws = wb[name]
        assert ws.max_row >= 1


def test_all_data_cells_have_text_format(pipeline):
    """데이터 셀(B5 등)이 '@' 텍스트 포맷이어야 '-' 시작 텍스트가 수식 해석 안 됨."""
    wb, parsed = pipeline
    for e in parsed.executives:
        ws = wb[_sheet_name(e.position)]
        assert ws["B5"].number_format == "@"
        assert ws["B6"].number_format == "@"
