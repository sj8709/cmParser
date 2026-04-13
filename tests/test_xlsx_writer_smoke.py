"""xlsx_writer 스모크 — 템플릿 복제·행 조정·기본 채움 검증."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from chaekmu_parser.extractors.docx_extractor import DocxExtractor
from chaekmu_parser.normalizer import normalize
from chaekmu_parser.xlsx_writer import write

ROOT = Path(__file__).parent.parent
FIXTURE = ROOT / "fixtures/ibk/input.docx"
TEMPLATE = ROOT / "templates/chaekmu_template.xlsx"
pytestmark = pytest.mark.skipif(
    not (FIXTURE.exists() and TEMPLATE.exists()),
    reason="fixture/template not present",
)


@pytest.fixture(scope="module")
def output_wb(tmp_path_factory):
    parsed = normalize(DocxExtractor().extract(FIXTURE))
    out = tmp_path_factory.mktemp("xlsx") / "ibk_out.xlsx"
    write(parsed, TEMPLATE, out)
    return load_workbook(str(out)), parsed


def test_setting_and_template_sheets_preserved(output_wb):
    wb, _ = output_wb
    assert "설정" in wb.sheetnames
    assert "임원_템플릿" in wb.sheetnames


def test_nine_exec_sheets_created(output_wb):
    wb, _ = output_wb
    exec_sheets = [n for n in wb.sheetnames if n not in ("설정", "임원_템플릿")]
    assert len(exec_sheets) == 9


def test_ceo_sheet_basic_fields(output_wb):
    wb, parsed = output_wb
    ceo = parsed.executives[0]
    sheet_name = ceo.position.replace("\n", ", ")[:31]
    ws = wb[sheet_name]
    assert "대표이사" in str(ws["B5"].value)
    assert ws["B6"].value == "전병성"
    assert ws["B9"].value == "Y"


def test_marketing_committee_rows_expanded_to_seven(output_wb):
    wb, parsed = output_wb
    marketing = next(e for e in parsed.executives if "마케팅" in e.position)
    sheet_name = marketing.position.replace("\n", ", ")[:31]
    ws = wb[sheet_name]
    # 회의체 데이터 시작은 row 15, 7개 → 15~21
    for r in range(15, 22):
        assert ws.cell(row=r, column=1).value, f"row {r} A열 비어있음"
    # row 22는 §1 각주
    v = ws.cell(row=22, column=1).value or ""
    assert "각주" in v or "📎" in v


def _oblig_start(exec_):
    """조정 후 oblig start = 37 + delta_committee + delta_resp."""
    dc = (len(exec_.committees) or 1) - 4
    dr = max(len(exec_.responsibilities), 1) - 5
    return 37 + dc + dr


def test_ceo_obligations_all_고유(output_wb):
    wb, parsed = output_wb
    ceo = parsed.executives[0]
    ws = wb[ceo.position.replace("\n", ", ")[:31]]
    oblig_start = _oblig_start(ceo)
    for idx in range(len(ceo.obligations)):
        row = oblig_start + idx
        assert ws.cell(row=row, column=1).value == "고유 책무", (
            f"CEO row {row} A열이 '고유 책무'가 아님"
        )


def test_non_ceo_last_obligation_공통(output_wb):
    wb, parsed = output_wb
    marketing = next(e for e in parsed.executives if "마케팅" in e.position)
    ws = wb[marketing.position.replace("\n", ", ")[:31]]
    oblig_start = _oblig_start(marketing)
    last_row = oblig_start + len(marketing.obligations) - 1
    assert ws.cell(row=last_row, column=1).value == "공통 책무"


def test_footnotes_preserved(output_wb):
    """각주 B19/B32/B46이 delta만큼 밀린 위치에 원문 보존."""
    wb, parsed = output_wb
    marketing = next(e for e in parsed.executives if "마케팅" in e.position)
    ws = wb[marketing.position.replace("\n", ", ")[:31]]
    dc = (len(marketing.committees) or 1) - 4
    dr = max(len(marketing.responsibilities), 1) - 5
    footer1 = 19 + dc
    footer2 = 32 + dc + dr
    footer3 = 46 + dc + dr + (max(len(marketing.obligations), 1) - 9)
    assert "주" in str(ws.cell(footer1, 2).value or ""), "§1 각주 본문 누락"
    assert "주" in str(ws.cell(footer2, 2).value or ""), "§2 각주 본문 누락"
    assert "주" in str(ws.cell(footer3, 2).value or ""), "§3 각주 본문 누락"


def test_resp_summary_placed_correctly(output_wb):
    """책무 개요/배분일자가 회의체 delta만큼 밀린 위치에 채워짐."""
    wb, parsed = output_wb
    marketing = next(e for e in parsed.executives if "마케팅" in e.position)
    ws = wb[marketing.position.replace("\n", ", ")[:31]]
    dc = (len(marketing.committees) or 1) - 4
    summary_row = 22 + dc
    assign_row = 23 + dc
    assert ws.cell(summary_row, 2).value == marketing.responsibility_summary
    assert ws.cell(assign_row, 2).value == marketing.assign_date


def test_text_format_applied(output_wb):
    wb, parsed = output_wb
    ceo = parsed.executives[0]
    sheet_name = ceo.position.replace("\n", ", ")[:31]
    ws = wb[sheet_name]
    assert ws["B5"].number_format == "@"
