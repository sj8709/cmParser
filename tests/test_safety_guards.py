"""안전장치 회귀 — S2(xlsx delete guard), S3/S4/S5는 GUI 단위 테스트."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from chaekmu_parser.extractors.docx_extractor import DocxExtractor
from chaekmu_parser.normalizer import normalize
from chaekmu_parser.xlsx_writer import _adjust_block, write

ROOT = Path(__file__).parent.parent
FIXTURE = ROOT / "fixtures/ibk/input.docx"
TEMPLATE = ROOT / "templates/chaekmu_template.xlsx"


# -----------------------------------------------------------------
# S2: xlsx_writer _adjust_block의 delete_at 가드
# -----------------------------------------------------------------
def test_adjust_block_raises_on_invalid_negative_delta():
    """last_data_row=5, footer_row=6에서 delta=-10 이면 delete_at=-4 → 오류."""
    wb = load_workbook(str(TEMPLATE))
    ws = wb["임원_템플릿"]
    with pytest.raises(ValueError, match="delete_rows 좌표 이상"):
        _adjust_block(ws, footer_row=6, last_data_row=5, delta=-10)


def test_adjust_block_accepts_valid_delete():
    """정상 범위 delete는 예외 없음."""
    wb = load_workbook(str(TEMPLATE))
    ws = wb["임원_템플릿"]
    # base 9(행 37~45), footer 46, delta=-4 → delete_at=42, 4행 삭제 (valid)
    _adjust_block(ws, footer_row=46, last_data_row=45, delta=-4)


def test_write_full_pipeline_still_works_after_guard(tmp_path):
    """IBK 실측으로 가드 추가 후에도 파이프라인이 여전히 작동."""
    if not FIXTURE.exists():
        pytest.skip("IBK fixture 없음")
    parsed = normalize(DocxExtractor().extract(FIXTURE))
    out = tmp_path / "out.xlsx"
    write(parsed, TEMPLATE, out)
    assert out.exists()
    wb = load_workbook(str(out))
    # 임원 시트 9개 + 설정 + 임원_템플릿 = 11
    assert len(wb.sheetnames) == 11


# -----------------------------------------------------------------
# S4: _open_in_file_manager 플랫폼 가드 (맥/리눅스에서 크래시 안 함)
# -----------------------------------------------------------------
def test_open_in_file_manager_no_crash_on_missing_path(monkeypatch):
    """존재하지 않는 경로 전달 시에도 subprocess/startfile이 예외 나면 로깅만."""
    from chaekmu_parser_gui.app import _open_in_file_manager

    class FakeWindow:
        def __init__(self):
            self.logged: list[str] = []

        def _log_line(self, t: str) -> None:
            self.logged.append(t)

    fake = FakeWindow()
    # 가짜 path — subprocess/startfile이 실패해도 함수 자체는 return
    monkeypatch.setattr("sys.platform", "win32")

    def fake_startfile(_path):
        raise FileNotFoundError("stub")

    import os as _os

    monkeypatch.setattr(_os, "startfile", fake_startfile, raising=False)

    _open_in_file_manager(Path("C:/nope"), fake)  # type: ignore[arg-type]
    assert any("열기 실패" in line for line in fake.logged)


# -----------------------------------------------------------------
# S5: _POLL_MAX_TICKS 상수 sanity
# -----------------------------------------------------------------
def test_poll_max_ticks_is_reasonable():
    from chaekmu_parser_gui.app import _POLL_INTERVAL_MS, _POLL_MAX_TICKS
    total_ms = _POLL_INTERVAL_MS * _POLL_MAX_TICKS
    assert 60_000 <= total_ms <= 600_000, f"폴링 타임아웃 {total_ms}ms은 1~10분 범위여야 함"
