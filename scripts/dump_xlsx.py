"""템플릿/정답 XLSX 구조 덤프 — Step A 분석용."""
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

def dump(path: Path, max_rows: int = 80):
    print(f"\n=== XLSX: {path.name} ===")
    wb = load_workbook(str(path), data_only=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: '{sheet_name}' dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column} ---")
        print(f"Merged ranges ({len(ws.merged_cells.ranges)}):")
        for mr in list(ws.merged_cells.ranges)[:30]:
            print(f"  {mr}")

        # 셀 내용
        print("Cell contents (non-empty):")
        shown = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows)):
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value).replace("\n", " | ")[:120]
                print(f"  {cell.coordinate}: {val}")
                shown += 1
        print(f"  (total non-empty shown: {shown})")


if __name__ == "__main__":
    root = Path(__file__).parent.parent
    target = sys.argv[1] if len(sys.argv) > 1 else "both"
    if target in ("template", "both"):
        dump(root / "templates/chaekmu_template.xlsx")
    if target in ("expected", "both"):
        dump(root / "fixtures/ibk/expected.xlsx", max_rows=200)
